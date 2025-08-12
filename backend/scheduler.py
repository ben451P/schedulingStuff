import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from backend.utils import minutes_to_time, military_to_normal, time_to_minutes
from backend.guard import Guard
from backend.station import Station
import io

class Scheduler:
    def __init__(self, start, end, lunch_start, lunch_end, rotation_cycle, importance_order, coverage_times, shifts):
        self.shifts = shifts
        self.guards = self.schedule_to_class()
        self.complete_schedule = False

        self.start = time_to_minutes(start)
        self.end = time_to_minutes(end)
        self.lunch_start = time_to_minutes(lunch_start)
        self.lunch_end = time_to_minutes(lunch_end)

        self.rotation_cycle = rotation_cycle
        self.station_importance_descending = importance_order
        self.coverage_times = coverage_times

        #schedule rows are time, cols are stations, in order of importance not actual rotation thing
        self.schedule = [[-1 for _ in self.rotation_cycle] for _ in range(self.start,self.end,15)]

        self.station_map = {i:Station(i,self.coverage_times[i]) for i in self.rotation_cycle}

    def schedule_to_class(self):
        guards = []
        for name, start, end in self.shifts:
            guard = Guard(name, start, end)
            for g in range(len(guards)):
                if guards[g].name == guard.name:
                    guards[g].start_time = min(guards[g].start_time, guard.start_time)
                    guards[g].end_time = max(guards[g].end_time, guard.end_time)
                    guards[g].determine_if_lunch_break()
                    break
            else:
                guards.append(guard)
        return guards
    
    def manually_override_lunches(self, lunches):
        for i in range(len(self.guards)):
            if not self.guards[i].start_time == self.guards[i].end_time:
                self.guards[i].lunch_break = lunches[i]


    def available_guards(self, time_str):
        time = time_to_minutes(time_str)
        availability = [int(g.is_available_at(time)) for g in self.guards]
        return availability, availability.count(True)
    
    def needed_stations(self,time):
        needed = [0 for _ in range(len(self.station_importance_descending))]
        for j, i in enumerate(self.station_importance_descending):
            station = self.station_map[i]
            for t in range(time,time + 60, 15):
                if station.should_be_open_at(t):
                    needed[j] = 1
                    break
                
        return needed, needed.count(1)

    def schedule_lunches(self):
    
        period_start, period_end = self.lunch_start, self.lunch_end
        drop = 0
        finished_scheduling = False

        needed_lunch_breaks = [guard.lunch_break for guard in self.guards]

        while not finished_scheduling:
            check = needed_lunch_breaks.copy()
            temp_guards_on_break = []
            for time in range(period_start,period_end, 15):
                avail_guards, num_avail_guards = self.available_guards(minutes_to_time(time))
                needed_stats, num_needed_stats = self.needed_stations(time)

                difference = num_needed_stats - num_avail_guards - drop + len(temp_guards_on_break)

                #need to correct for assumption that least important station is gonna go
                while difference < 0:
                    #give lunch breaks while there is a difference
                    if True in check:
                        index = check.index(True)
                        check[index] = time
                        temp_guards_on_break.append(4)
                    difference += 1
                #increment for looop
                pointer = 0
                while pointer < len(temp_guards_on_break):
                    temp_guards_on_break[pointer] -= 1
                    if temp_guards_on_break[pointer] == 0:
                        temp_guards_on_break.pop(pointer)
                    else:pointer += 1

            if True not in check:
                finished_scheduling = True
            # check if everyones break is scheduled
            drop += 1
        
        #updating the actual info
        for i in range(len(self.guards)):
            if check[i]:
                self.guards[i].lunch_break_start = check[i]
                self.guards[i].lunch_break_end = check[i] + 60

    def add_fodder_station(self, index):
        count = 1
        for i in self.rotation_cycle:
            if "Standby" in i:
                count += 1
        station_name = "Standby" + str(count)

        self.station_importance_descending.insert(0,station_name)
        self.rotation_cycle.append(station_name)

        for i in range(index):
            self.schedule[i].append(-1)

    def create_base_schedule(self):
        time = self.start

        prev_availability, prev_num_avail = self.available_guards(minutes_to_time(time))

        prev_state = []
        for i,j in enumerate(prev_availability):
            if j:prev_state.append(i)

        for row in range(len(self.schedule)):
            availability,num_avail = self.available_guards(minutes_to_time(time))

            if num_avail > len(self.schedule[row]):
                self.add_fodder_station(row)

            #does default rotation unless new, less or different guards than before
            new_state = prev_state.copy()
            if prev_availability != availability:
                for i in range(len(availability)):
                    #if guard is leaving mark station as unattended
                    if not availability[i] and prev_availability[i]:
                        new_state[new_state.index(i)] = -1
                #if different guards, find an unattended station and man it
                for i in range(len(availability)):
                    if availability[i] and not prev_availability[i]:
                        if -1 in new_state:
                            new_state[new_state.index(-1)] = i
                
            #if more guards than before, open next most impotant station
            if num_avail > prev_num_avail:
                for guard_num in range(len(availability)):
                    if availability[guard_num] and guard_num not in new_state:
                        new_state.append(guard_num)

            #if less guards, shift so that least important station is unattended
            elif num_avail < prev_num_avail:
                while -1 in new_state:
                    index = new_state.index(-1)
                    temp = new_state[index]
                    new_state[index] = new_state[-1]
                    new_state.pop(-1)

            #this is a rotation #####
            temp = new_state.copy() + [-1] * (len(self.rotation_cycle) - len(new_state))
            importance_index = {name: i for i, name in enumerate(self.station_importance_descending[::-1])}
            reordered = [temp[importance_index[station]] for station in self.rotation_cycle]
            while reordered and reordered[-1] == -1:
                reordered.pop(-1)

            front_negs = 0
            while reordered and reordered[0] == -1:
                reordered.pop(0)
                front_negs += 1
            temp = reordered[-1]
            gap = 1
            pointer = len(reordered) - 1
            # Remove temp from its current position before rotation
            while pointer > 0:
                if reordered[pointer - gap] == -1:
                    gap += 1
                else:
                    reordered[pointer] = reordered[pointer - gap]
                    pointer -= gap
                    gap = 1
            reordered[0] = temp

            reordered = [-1] * front_negs + reordered

            temp = reordered.copy() + [-1] * (len(self.rotation_cycle) - len(reordered))
            cycle_index = {name: i for i, name in enumerate(self.rotation_cycle)}
            original_order = [temp[cycle_index[station]] for station in self.station_importance_descending[::-1]]
            while original_order and original_order[-1] == -1:
                original_order.pop(-1)
            
            new_state = original_order.copy()
            #up to here ####

            self.schedule[row] = new_state + [-1] * (len(self.rotation_cycle) - len(new_state))

            time += 15
            prev_availability = availability.copy()
            prev_num_avail = num_avail
            prev_state = new_state.copy()
        
        for i in range(len(self.schedule)):
            for j in range(len(self.schedule[i])):
                if self.schedule[i][j] != -1:
                    self.schedule[i][j] += 1
                    

    def convert_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # build time labels
        times = [minutes_to_time(t) for t in range(self.start, self.end, 15)]
        times = [military_to_normal(t) for t in times]

        # build dataframe
        df = pd.DataFrame(
            self.schedule,
            index=times,
            columns=self.station_importance_descending[::-1]
        )
        df = df[self.rotation_cycle]
        df = df.T

        # headers
        ws['A1'] = 'Time'
        for col_idx, time in enumerate(times, start=2):
            ws.cell(row=1, column=col_idx, value=time)

        # schedule table
        for row_idx, station in enumerate(self.rotation_cycle, start=2):
            ws.cell(row=row_idx, column=1, value=station)
            for col_idx, time in enumerate(times, start=2):
                value = df.loc[station, time]
                ws.cell(row=row_idx, column=col_idx, value="" if value == -1 else value)

        # style anomalies
        anomaly_fill = self._apply_excel_styling(ws, len(times))
        anomaly_cells = self.detect_rotation_anomalies(df)
        for row, col in anomaly_cells:
            ws.cell(row=row, column=col).fill = anomaly_fill

        # lunch breaks table (3 rows below schedule)
        lunch_start_row = len(self.rotation_cycle) + 4
        ws.cell(row=lunch_start_row, column=1, value="Guard")
        ws.cell(row=lunch_start_row, column=2, value="Break Start")

        current_row = lunch_start_row + 1
        for guard in self.guards:
            if guard.lunch_break:
                ws.cell(row=current_row, column=1, value=guard.name)
                ws.cell(row=current_row, column=2, value=military_to_normal(minutes_to_time(guard.lunch_break_start)))
                current_row += 1

        # save to memory for Flask download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    
    def detect_rotation_anomalies(self, df):
        anomaly_cells = set()
        
        time_cols = df.columns.tolist()
        
        for time_idx in range(len(time_cols) - 1):
            current_time = time_cols[time_idx]
            next_time = time_cols[time_idx + 1]

            # guard -> station_idx
            current_staffed = {}  
            next_staffed = {}
            
            for station_idx, station in enumerate(self.rotation_cycle):
                current_guard = df.loc[station, current_time]
                next_guard = df.loc[station, next_time]
                
                if current_guard != "" and current_guard != -1:
                    current_staffed[current_guard] = station_idx
                if next_guard != "" and next_guard != -1:
                    next_staffed[next_guard] = station_idx
            
            for guard in current_staffed:
                if guard in next_staffed:
                    current_station_idx = current_staffed[guard]
                    next_station_idx = next_staffed[guard]
                    
                    expected_next_idx = None
                    
                    for offset in range(1, len(self.rotation_cycle)):
                        candidate_idx = (current_station_idx + offset) % len(self.rotation_cycle)
                        candidate_station = self.rotation_cycle[candidate_idx]
                        
                        if df.loc[candidate_station, next_time] != "" and df.loc[candidate_station, next_time] != -1:
                            expected_next_idx = candidate_idx
                            break
                    
                    if expected_next_idx is not None and next_station_idx != expected_next_idx:

                        anomaly_cells.add((current_station_idx + 2, time_idx + 2))
                        anomaly_cells.add((next_station_idx + 2, time_idx + 3))
        
        return anomaly_cells

    def _apply_excel_styling(self, ws, num_time_cols):
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        station_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        anomaly_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        header_font = Font(color="FFFFFF", bold=True, size=10)
        station_font = Font(bold=True, size=10)
        data_font = Font(size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for col in range(1, num_time_cols + 2):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        for row in range(2, len(self.rotation_cycle) + 2):
            station_cell = ws.cell(row=row, column=1)
            station_cell.fill = station_fill
            station_cell.font = station_font
            station_cell.alignment = center_alignment
            station_cell.border = thin_border
            
            for col in range(2, num_time_cols + 2):
                data_cell = ws.cell(row=row, column=col)
                data_cell.fill = data_fill
                data_cell.font = data_font
                data_cell.alignment = center_alignment
                data_cell.border = thin_border
        
        ws.column_dimensions['A'].width = 12
        for col_idx in range(2, num_time_cols + 2):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 6
        
        for row in range(1, len(self.rotation_cycle) + 2):
            ws.row_dimensions[row].height = 20
        
        return anomaly_fill