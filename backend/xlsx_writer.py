import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io

from .scheduler import Scheduler
from .utils import minutes_to_time, military_to_normal


class XLSXWriter:
    def __init__(self, scheduler: Scheduler):
        self.scheduler = scheduler

    def convert_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # build time labels
        times = [minutes_to_time(t) for t in range(self.scheduler.start, self.scheduler.end, 15)]
        times = [military_to_normal(t) for t in times]

        # build dataframe
        df = pd.DataFrame(
            self.scheduler.schedule,
            index=times,
            columns=self.scheduler.station_importance_descending[::-1]
        )
        df = df[self.scheduler.rotation_cycle]
        df = df.T

        # headers
        ws['A1'] = 'Time'
        for col_idx, time in enumerate(times, start=2):
            ws.cell(row=1, column=col_idx, value=time)

        # schedule table
        for row_idx, station in enumerate(self.scheduler.rotation_cycle, start=2):
            ws.cell(row=row_idx, column=1, value=station)
            for col_idx, time in enumerate(times, start=2):
                value = df.loc[station, time]
                ws.cell(row=row_idx, column=col_idx, value="" if value == -1 else value)

        # style anomalies
        anomaly_fill = self._apply_excel_styling(ws, len(times))
        anomaly_cells = self.detect_rotation_anomalies(df)
        for row, col in anomaly_cells:
            ws.cell(row=row, column=col).fill = anomaly_fill

        # lunch breaks table (3 rows below schedule)
        lunch_start_row = len(self.scheduler.rotation_cycle) + 4
        ws.cell(row=lunch_start_row, column=1, value="Guard")
        ws.cell(row=lunch_start_row, column=2, value="Break Start")

        current_row = lunch_start_row + 1
        for guard in self.scheduler.guards:
            if guard.lunch_break:
                ws.cell(row=current_row, column=1, value=guard.name)
                ws.cell(row=current_row, column=2, value=military_to_normal(minutes_to_time(guard.lunch_break_start)))
                current_row += 1

        # save to memory for Flask download
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    
    def detect_rotation_anomalies(self, df):
        anomaly_cells = set()
        
        time_cols = df.columns.tolist()
        
        for time_idx in range(len(time_cols) - 1):
            current_time = time_cols[time_idx]
            next_time = time_cols[time_idx + 1]

            # guard -> station_idx
            current_staffed = {}  
            next_staffed = {}
            
            for station_idx, station in enumerate(self.scheduler.rotation_cycle):
                current_guard = df.loc[station, current_time]
                next_guard = df.loc[station, next_time]
                
                if current_guard != "" and current_guard != -1:
                    current_staffed[current_guard] = station_idx
                if next_guard != "" and next_guard != -1:
                    next_staffed[next_guard] = station_idx
            
            for guard in current_staffed:
                if guard in next_staffed:
                    current_station_idx = current_staffed[guard]
                    next_station_idx = next_staffed[guard]
                    
                    expected_next_idx = None
                    
                    for offset in range(1, len(self.scheduler.rotation_cycle)):
                        candidate_idx = (current_station_idx + offset) % len(self.scheduler.rotation_cycle)
                        candidate_station = self.scheduler.rotation_cycle[candidate_idx]
                        
                        if df.loc[candidate_station, next_time] != "" and df.loc[candidate_station, next_time] != -1:
                            expected_next_idx = candidate_idx
                            break
                    
                    if expected_next_idx is not None and next_station_idx != expected_next_idx:

                        anomaly_cells.add((current_station_idx + 2, time_idx + 2))
                        anomaly_cells.add((next_station_idx + 2, time_idx + 3))
        
        return anomaly_cells

    def _apply_excel_styling(self, ws, num_time_cols):
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        station_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        anomaly_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        header_font = Font(color="FFFFFF", bold=True, size=10)
        station_font = Font(bold=True, size=10)
        data_font = Font(size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for col in range(1, num_time_cols + 2):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        for row in range(2, len(self.scheduler.rotation_cycle) + 2):
            station_cell = ws.cell(row=row, column=1)
            station_cell.fill = station_fill
            station_cell.font = station_font
            station_cell.alignment = center_alignment
            station_cell.border = thin_border
            
            for col in range(2, num_time_cols + 2):
                data_cell = ws.cell(row=row, column=col)
                data_cell.fill = data_fill
                data_cell.font = data_font
                data_cell.alignment = center_alignment
                data_cell.border = thin_border
        
        ws.column_dimensions['A'].width = 12
        for col_idx in range(2, num_time_cols + 2):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 6
        
        for row in range(1, len(self.scheduler.rotation_cycle) + 2):
            ws.row_dimensions[row].height = 20
        
        return anomaly_fill