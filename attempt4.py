from datetime import datetime
import numpy as np
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows



class Station:
    def __init__(self, name, times_when_open: list):
        self.name = name
        self.times_when_open = times_when_open #list with tuples (start_time,end_time)

    def __repr__(self):
        return self.name
    
    def should_be_open_at(self,time):
        for item in self.times_when_open:
            if time_to_minutes(item[0]) <= time < time_to_minutes(item[1]):
                return True
        return False

ROTATION_CYCLE = [
    "Kiddie", "Dive", "Main", "Break", "First Aid", "Slide",
    "Main2", "Rover", "Lap", "See Manager", "Bathroom Break"
]
STATION_IMPORTANCE_DESCENDING = [
    "Bathroom Break", "Rover", "Main2", "See Manager", "Slide",
    "Kiddie", "First Aid", "Dive", "Lap", "Main", "Break"
]

STATION_COVERAGE_TIMES = {
    i:[("11:00", "20:00")] for i in ROTATION_CYCLE
}

STATION_MAP = {
    i:Station(i,STATION_COVERAGE_TIMES[i]) for i in ROTATION_CYCLE
}

shifts = {
    "Guard A": ("09:45", "15:30"),
    "Guard B": ("09:45", "15:30"),
    "Guard C": ("10:30", "16:00"),
    "Guard D": ("10:30", "16:00"),
    "Guard E": ("11:00", "20:00"),
    "Guard F": ("11:00", "20:00"),
    "Guard G": ("11:00", "20:00"),
    "Guard H": ("11:00", "20:00"),
    "Guard I": ("13:00", "19:00"),
    "Guard J": ("14:00", "20:00"),
    "Guard K": ("14:00", "20:00"),
    "Guard L": ("14:00", "20:00"),
    "Guard M": ("15:30", "20:00"),
}

def time_to_minutes(t: str) -> int:
    h, m = map(int, t.split(":"))
    return h * 60 + m

def minutes_to_time(minutes: int) -> str:
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"

ACCEPTABLE_LUNCH_START_TIME_RANGE = (time_to_minutes("13:00"), time_to_minutes("16:00"))

class Guard:
    def __init__(self, name, start, end):
        self.name = name
        self.start_time = time_to_minutes(start)
        self.end_time = time_to_minutes(end)
        self.lunch_break = self.end_time - self.start_time > 360
        self.lunch_break_start = 0
        self.lunch_break_end = 0

    def __repr__(self):
        return self.name

    def is_available_at(self, time: int) -> bool:
        on_shift = self.start_time <= time < self.end_time
        on_break = self.lunch_break_start <= time < self.lunch_break_end
        return on_shift and not on_break

class Scheduler:
    def __init__(self, shifts):
        self.shifts = shifts
        self.guards = self.schedule_to_class()
        self.complete_schedule = False
        self.start = time_to_minutes("11:00")
        self.end = time_to_minutes("19:30")
        #schedule rows are time, cols are stations, in order of importance not actual rotation thing
        self.schedule = [[-1 for i in ROTATION_CYCLE] for _ in range(self.start,self.end,15)]

    def schedule_to_class(self):
        guards = []
        for name, (start, end) in self.shifts.items():
            guard = Guard(name, start, end)
            for g in range(len(guards)):
                if guards[g].name == guard.name:
                    guards[g].start_time = min(guards[g].start_time, guard.start_time)
                    guards[g].end_time = max(guards[g].end_time, guard.end_time)
                    break
            else:
                guards.append(guard)
        return guards


    def exchange_numbers(self,num1, num2, time):
        up_to = (time - self.start_time) / 15
        for j in range(up_to):
            for i in range(len(self.schedule[j])):
                if self.schedule[j][i] == num1:
                    self.schedule[j][i] = num2
                elif self.schedule[j][i] == num2:
                    self.schedule[j][i] = num1
        return self.schedule

    def available_guards(self, time_str):
        time = time_to_minutes(time_str)
        availability = [int(g.is_available_at(time)) for g in self.guards]
        return availability, availability.count(True)
    
    def needed_stations(self,time):
        needed = [0 for _ in range(len(STATION_IMPORTANCE_DESCENDING))]
        for j, i in enumerate(STATION_IMPORTANCE_DESCENDING):
            station = STATION_MAP[i]
            for t in range(time,time + 60, 15):
                if station.should_be_open_at(t):
                    needed[j] = 1
                    break
                
        return needed, needed.count(1)

    def schedule_lunches(self):
    
        period_start, period_end = ACCEPTABLE_LUNCH_START_TIME_RANGE
        drop = 0
        finished_scheduling = False

        needed_lunch_breaks = [guard.lunch_break for guard in self.guards]

        

        while not finished_scheduling:
            check = needed_lunch_breaks.copy()
            temp_guards_on_break = []
            for time in range(period_start,period_end, 15):
                avail_guards, num_avail_guards = self.available_guards(minutes_to_time(time))
                needed_stats, num_needed_stats = self.needed_stations(time)

                difference = num_needed_stats - num_avail_guards - drop + len(temp_guards_on_break)

                #need to correct for assumption that least important station is gonna go
                while difference < 0:
                    #give lunch breaks while there is a difference
                    if True in check:
                        index = check.index(True)
                        check[index] = time
                        temp_guards_on_break.append(4)
                    difference += 1
                #increment for looop
                pointer = 0
                while pointer < len(temp_guards_on_break):
                    temp_guards_on_break[pointer] -= 1
                    if temp_guards_on_break[pointer] == 0:
                        temp_guards_on_break.pop(pointer)
                    else:pointer += 1

            if True not in check:
                finished_scheduling = True
            # check if everyones break is scheduled
            drop += 1
        
        #updating the actual info
        for i in range(len(self.guards)):
            if check[i]:
                self.guards[i].lunch_break_start = check[i]
                self.guards[i].lunch_break_end = check[i] + 60

    def create_base_schedule(self):
        time = time_to_minutes("11:00")

        prev_availability, prev_num_avail = self.available_guards(minutes_to_time(time))

        prev_state = []
        for i,j in enumerate(prev_availability):
            if j:prev_state.append(i)

        for row in range(len(self.schedule)):
            availability,num_avail = self.available_guards(minutes_to_time(time))

            #this is a rotation #####
            new_state = prev_state.copy()

            temp = new_state.copy() + [-1] * (len(ROTATION_CYCLE) - len(new_state))
            importance_index = {name: i for i, name in enumerate(STATION_IMPORTANCE_DESCENDING[::-1])}
            reordered = [temp[importance_index[station]] for station in ROTATION_CYCLE]
            while reordered and reordered[-1] == -1:
                reordered.pop(-1)

            temp = reordered[-1]
            gap = 1
            pointer = len(reordered) - 1
            while pointer > 0:
                if reordered[pointer - gap] == -1:
                    gap += 1
                else:
                    reordered[pointer] = reordered[pointer - gap]
                    pointer -= gap
                    gap = 1
            reordered[0] = temp

            temp = reordered.copy() + [-1] * (len(ROTATION_CYCLE) - len(reordered))
            cycle_index = {name: i for i, name in enumerate(ROTATION_CYCLE)}
            original_order = [temp[cycle_index[station]] for station in STATION_IMPORTANCE_DESCENDING[::-1]]
            while original_order and original_order[-1] == -1:
                original_order.pop(-1)
            
            new_state = original_order.copy()
            #up to here ####

            if prev_availability != availability:
                for i in range(len(availability)):
                    
                    if not availability[i] and prev_availability[i]:
                        new_state[new_state.index(i)] = -1
                for i in range(len(availability)):
                    if availability[i] and not prev_availability[i]:
                        if -1 in new_state:
                            new_state[new_state.index(-1)] = i
                                    
            if num_avail > prev_num_avail:
                for guard_num in range(len(availability)):
                    if availability[guard_num] and guard_num not in new_state:
                        new_state.append(guard_num)

            elif num_avail < prev_num_avail:
                while -1 in new_state:
                    index = new_state.index(-1)
                    temp = new_state[index]
                    new_state[index] = new_state[-1]
                    new_state.pop(-1)


            self.schedule[row] = new_state + [-1] * (len(ROTATION_CYCLE) - len(new_state))

            time += 15
            prev_availability = availability.copy()
            prev_num_avail = num_avail
            prev_state = new_state.copy()
                    

    def convert_to_excel(self):
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Generate time headers (15-minute intervals)
        times = [minutes_to_time(t) for t in range(self.start, self.end, 15)]
        
        # Create DataFrame for easier manipulation
        df = pd.DataFrame(
            self.schedule,
            index=times,
            columns=STATION_IMPORTANCE_DESCENDING[::-1]
        )
        df = df[ROTATION_CYCLE]  # Reorder columns to match rotation cycle
        df = df.T  # Transpose so stations are rows
        
        # Write headers
        ws['A1'] = 'Time'
        for col_idx, time in enumerate(times, start=2):
            ws.cell(row=1, column=col_idx, value=time)
        
        # Write station names and data
        for row_idx, station in enumerate(ROTATION_CYCLE, start=2):
            ws.cell(row=row_idx, column=1, value=station)
            for col_idx, time in enumerate(times, start=2):
                value = df.loc[station, time]
                if value == -1:
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply styling and get anomaly fill
        anomaly_fill = self._apply_excel_styling(ws, len(times))
        
        # Detect and highlight rotation anomalies
        anomaly_cells = self.detect_rotation_anomalies(df)
        for row, col in anomaly_cells:
            ws.cell(row=row, column=col).fill = anomaly_fill
        
        # Save workbook
        wb.save("schedule.xlsx")
        print("Excel file written: schedule.xlsx")
    
    def detect_rotation_anomalies(self, df):
        """
        Detect when guards don't follow normal rotation pattern.
        Skipping unstaffed stations is normal. Only flag when a guard jumps
        to a station that should have been skipped or doesn't follow the expected
        sequence through staffed stations.
        Returns set of (row, col) tuples for cells to highlight in gray.
        """
        anomaly_cells = set()
        
        # Get list of time columns
        time_cols = df.columns.tolist()
        
        for time_idx in range(len(time_cols) - 1):
            current_time = time_cols[time_idx]
            next_time = time_cols[time_idx + 1]
            
            # Get staffed stations for current and next time
            current_staffed = {}  # guard -> station_idx
            next_staffed = {}     # guard -> station_idx
            
            for station_idx, station in enumerate(ROTATION_CYCLE):
                current_guard = df.loc[station, current_time]
                next_guard = df.loc[station, next_time]
                
                if current_guard != "" and current_guard != -1:
                    current_staffed[current_guard] = station_idx
                if next_guard != "" and next_guard != -1:
                    next_staffed[next_guard] = station_idx
            
            # Check each guard's movement
            for guard in current_staffed:
                if guard in next_staffed:
                    current_station_idx = current_staffed[guard]
                    next_station_idx = next_staffed[guard]
                    
                    # Find the next staffed station in rotation order
                    expected_next_idx = None
                    
                    # Look for next staffed station starting from current position + 1
                    for offset in range(1, len(ROTATION_CYCLE)):
                        candidate_idx = (current_station_idx + offset) % len(ROTATION_CYCLE)
                        candidate_station = ROTATION_CYCLE[candidate_idx]
                        
                        # Check if this station is staffed in the next time slot
                        if df.loc[candidate_station, next_time] != "" and df.loc[candidate_station, next_time] != -1:
                            expected_next_idx = candidate_idx
                            break
                    
                    # If we found an expected next station and guard didn't go there
                    if expected_next_idx is not None and next_station_idx != expected_next_idx:
                        # Mark current position as anomaly
                        anomaly_cells.add((current_station_idx + 2, time_idx + 2))
                        # Mark next position as anomaly  
                        anomaly_cells.add((next_station_idx + 2, time_idx + 3))
        
        return anomaly_cells

    def _apply_excel_styling(self, ws, num_time_cols):
        # Define colors and styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        station_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        anomaly_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gray for anomalies
        
        header_font = Font(color="FFFFFF", bold=True, size=10)
        station_font = Font(bold=True, size=10)
        data_font = Font(size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Style header row (time row)
        for col in range(1, num_time_cols + 2):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Style station column and data
        for row in range(2, len(ROTATION_CYCLE) + 2):
            # Station name column
            station_cell = ws.cell(row=row, column=1)
            station_cell.fill = station_fill
            station_cell.font = station_font
            station_cell.alignment = center_alignment
            station_cell.border = thin_border
            
            # Data columns
            for col in range(2, num_time_cols + 2):
                data_cell = ws.cell(row=row, column=col)
                data_cell.fill = data_fill
                data_cell.font = data_font
                data_cell.alignment = center_alignment
                data_cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        for col_idx in range(2, num_time_cols + 2):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 6
        
        # Set row height
        for row in range(1, len(ROTATION_CYCLE) + 2):
            ws.row_dimensions[row].height = 20
        
        return anomaly_fill

scheduler = Scheduler(shifts)
scheduler.schedule_lunches()
scheduler.create_base_schedule()
scheduler.convert_to_excel()




